#!/usr/bin/env python3
"""
export_ca_alerts_csv.py

Fast script to stream and export all CA asset alerts (1st Oct 2025 to till date)
directly to CSV at /opt/audit_report/CA_Alert_Report_2025_10_01_to_till_Date.csv
and /opt/audit_report/File-upload/CA_Alert_Report_2025_10_01_to_till_Date.csv
"""

import csv
import json
import os
import re
import shutil
import sys
from datetime import datetime, timezone
from typing import Dict, List, Any, Optional, Tuple

import requests
import openpyxl

PROM_URL = "http://localhost:9090"
LOG_FILE = "/var/log/prometheus/alertmanager_events.log"
INVENTORY_FILE = "/opt/audit_report/File-upload/SYSTEM_INVENTORY_REPORT.xlsx"
OUTPUT_DIR = "/opt/audit_report"
UPLOAD_DIR = "/opt/audit_report/File-upload"
OUTPUT_FILENAME = "CA_Alert_Report_2025_10_01_to_till_Date.csv"

START_DATE = datetime(2025, 10, 1, 0, 0, 0, tzinfo=timezone.utc)
END_DATE = datetime(2026, 8, 21, 23, 59, 59, tzinfo=timezone.utc)

HEADERS = [
    "Date",
    "Year",
    "Month",
    "Day",
    "Weekday",
    "Alert Name",
    "Asset",
    "Instance",
    "Job",
    "Group",
    "Severity",
    "Vital",
    "CPU Core",
    "Memory Total",
    "Total Disk Size",
    "Volume",
    "Free",
    "Used",
    "Summary",
    "Description"
]


def prom_query(prom_url: str, query: str) -> List[Dict[str, Any]]:
    try:
        r = requests.get(f"{prom_url}/api/v1/query", params={"query": query}, timeout=10)
        r.raise_for_status()
        return r.json().get("data", {}).get("result", [])
    except Exception:
        return []


def get_hardware_specs(prom_url: str) -> Dict[str, Dict[str, Any]]:
    specs: Dict[str, Dict[str, Any]] = {}
    for r in prom_query(prom_url, 'count(node_cpu_seconds_total{mode="idle"}) by (instance)'):
        inst = r.get("metric", {}).get("instance", "").split(":")[0]
        if inst:
            try:
                specs.setdefault(inst, {})["cpu_cores"] = int(r.get("value", [0, 0])[1])
            except (ValueError, TypeError, IndexError):
                pass

    for r in prom_query(prom_url, 'windows_cs_logical_processors'):
        inst = r.get("metric", {}).get("instance", "").split(":")[0]
        if inst and "cpu_cores" not in specs.get(inst, {}):
            try:
                specs.setdefault(inst, {})["cpu_cores"] = int(r.get("value", [0, 0])[1])
            except (ValueError, TypeError, IndexError):
                pass

    for r in prom_query(prom_url, 'node_memory_MemTotal_bytes'):
        inst = r.get("metric", {}).get("instance", "").split(":")[0]
        if inst:
            try:
                specs.setdefault(inst, {})["mem_total_bytes"] = float(r.get("value", [0, 0])[1])
            except (ValueError, TypeError, IndexError):
                pass

    for r in prom_query(prom_url, 'windows_cs_physical_memory_bytes'):
        inst = r.get("metric", {}).get("instance", "").split(":")[0]
        if inst and "mem_total_bytes" not in specs.get(inst, {}):
            try:
                specs.setdefault(inst, {})["mem_total_bytes"] = float(r.get("value", [0, 0])[1])
            except (ValueError, TypeError, IndexError):
                pass

    for r in prom_query(prom_url, 'node_filesystem_size_bytes'):
        inst = r.get("metric", {}).get("instance", "").split(":")[0]
        mp = r.get("metric", {}).get("mountpoint", "")
        if inst and mp:
            try:
                specs.setdefault(inst, {}).setdefault("disks", {})[mp] = float(r.get("value", [0, 0])[1])
            except (ValueError, TypeError, IndexError):
                pass

    for r in prom_query(prom_url, 'windows_logical_disk_size_bytes'):
        inst = r.get("metric", {}).get("instance", "").split(":")[0]
        vol = r.get("metric", {}).get("volume", "")
        if inst and vol:
            try:
                specs.setdefault(inst, {}).setdefault("disks", {})[vol] = float(r.get("value", [0, 0])[1])
            except (ValueError, TypeError, IndexError):
                pass

    return specs


def get_prometheus_assets_and_metadata(prom_url: str) -> Tuple[Dict[str, str], Dict[str, Dict[str, str]]]:
    assets_map: Dict[str, str] = {}
    meta_map: Dict[str, Dict[str, str]] = {}
    try:
        r = requests.get(f"{prom_url}/api/v1/targets", timeout=10)
        if r.status_code == 200:
            targets = r.json().get("data", {}).get("activeTargets", [])
            for t in targets:
                labels = t.get("labels", {})
                inst = labels.get("instance", "").split(":")[0]
                asset = labels.get("asset") or labels.get("Asset")
                if inst:
                    if asset:
                        assets_map[inst] = asset.strip()
                    meta_map[inst] = {
                        "job": labels.get("job", ""),
                        "group": labels.get("group", ""),
                        "company": labels.get("company", ""),
                    }
    except Exception:
        pass
    return assets_map, meta_map


def get_inventory_specs(inv_path: str) -> Dict[str, Dict[str, Any]]:
    inv: Dict[str, Dict[str, Any]] = {}
    if not os.path.exists(inv_path):
        return inv
    try:
        wb = openpyxl.load_workbook(inv_path, data_only=True)
        if "System Inventory Report" in wb.sheetnames:
            ws = wb["System Inventory Report"]
            for row in ws.iter_rows(values_only=True):
                if row and row[0] and row[0] != "IP Address" and not str(row[0]).startswith("SYSTEM") and not str(row[0]).startswith("Generated"):
                    ip = str(row[0]).strip()
                    inv[ip] = {
                        "hostname": str(row[1] or ""),
                        "job": str(row[2] or ""),
                        "cpu": str(row[3] or ""),
                        "memory": str(row[4] or ""),
                        "hdd": str(row[5] or ""),
                        "os": str(row[6] or ""),
                        "status": str(row[7] or ""),
                        "group": str(row[8] or ""),
                        "company": str(row[9] or "")
                    }
    except Exception as e:
        print(f"Warning: Could not read inventory: {e}", file=sys.stderr)
    return inv


def fmt_gb(v: Optional[float]) -> str:
    if not v or v <= 0:
        return "N/A"
    gb = v / (1024 ** 3)
    if gb >= 1000:
        return f"{gb / 1024:.2f} TB"
    return f"{gb:.2f} GB"


def extract_field(text: str, field_name: str) -> str:
    m = re.search(rf'{field_name}[:=]\s*([A-Za-z0-9\-_&\.]+)', text or "", re.IGNORECASE)
    return m.group(1).strip() if m else ""


def extract_group(alert_dict: Dict[str, Any], default_grp: str = "N/A") -> str:
    s = alert_dict.get("summary", "")
    d = alert_dict.get("description", "")
    full_text = f"{s}\n{d}"
    m = re.search(r'group[:=]\s*([A-Za-z0-9\-_&\.]+)', full_text, re.IGNORECASE)
    if m:
        return m.group(1).strip()
    return (
        alert_dict.get("group")
        or alert_dict.get("labels", {}).get("group")
        or default_grp
    )


def format_row(
    r: Dict[str, Any],
    hw_specs: Dict[str, Dict[str, Any]],
    inv_specs: Dict[str, Dict[str, Any]]
) -> List[Any]:
    ts: datetime = r["ts"]
    alert: str = r["alert"]
    inst: str = r["instance"]
    sev: str = r["severity"]
    desc: str = r["desc"]
    summ: str = r["summary"]
    full_text = f"{summ}\n{desc}"

    year_val = ts.year
    month_val = ts.strftime("%B")
    day_val = ts.day
    weekday_val = ts.strftime("%A")

    spec = hw_specs.get(inst, {})
    inv = inv_specs.get(inst, {})

    cpu_cores = spec.get("cpu_cores")
    if cpu_cores is None and inv.get("cpu"):
        cpu_str = inv["cpu"]
    elif cpu_cores is not None:
        cpu_str = f"{cpu_cores} Core"
    else:
        cpu_str = "N/A"

    mem_total_bytes = spec.get("mem_total_bytes")
    disks = spec.get("disks", {})

    m_vol = (
        re.search(r'(?:Drive|volume)[:=]?\s*([A-Z]:)', full_text, re.IGNORECASE)
        or re.search(r'(?:Mountpoint|mountpoint)[:=]?\s*([/\w\-_]+)', full_text, re.IGNORECASE)
    )
    volume = m_vol.group(1).upper() if m_vol and ":" in m_vol.group(1) else (m_vol.group(1) if m_vol else "N/A")

    total_disk_bytes = disks.get(volume)
    if not total_disk_bytes and volume != "N/A":
        for d_k, d_v in disks.items():
            if d_k.lower() == volume.lower():
                total_disk_bytes = d_v
                break
    if not total_disk_bytes and disks:
        total_disk_bytes = list(disks.values())[0]

    m_used = re.search(r'Used\s*=\s*([\d.]+)%', desc) or re.search(r'(\d+)%', sev)
    used_pct = float(m_used.group(1)) if m_used else None

    m_free = re.search(r'(?:Free Space|Available)\s*=\s*([0-9\.\s\wGBTB]+)', desc, re.IGNORECASE)
    free_str = m_free.group(1).strip() if m_free else "N/A"

    alert_lower = alert.lower()
    if "memory" in alert_lower or "mem" in alert_lower:
        vital_type = "Memory"
    elif "disk" in alert_lower or "space" in alert_lower or "mountpoint" in alert_lower:
        vital_type = "Disk"
    elif "cpu" in alert_lower:
        vital_type = "CPU"
    elif "swap" in alert_lower:
        vital_type = "Swap"
    elif "down" in alert_lower or "uptime" in alert_lower:
        vital_type = "System"
    elif "network" in alert_lower:
        vital_type = "Network"
    elif "service" in alert_lower or "port" in alert_lower:
        vital_type = "Service"
    else:
        vital_type = r.get("vital_extracted") or "General"

    mem_total_str = fmt_gb(mem_total_bytes) if mem_total_bytes else (inv.get("memory") or "N/A")
    disk_total_str = fmt_gb(total_disk_bytes) if total_disk_bytes else (inv.get("hdd") or "N/A")
    used_str = f"{used_pct:.2f}%" if used_pct is not None else "N/A"

    if vital_type == "Memory":
        if mem_total_bytes and used_pct is not None:
            tot_gb = mem_total_bytes / (1024 ** 3)
            used_gb = tot_gb * (used_pct / 100)
            free_gb = tot_gb - used_gb
            free_str = fmt_gb(free_gb * (1024 ** 3))
            used_str = fmt_gb(used_gb * (1024 ** 3))
    elif vital_type == "Disk":
        if total_disk_bytes and used_pct is not None:
            tot_gb = total_disk_bytes / (1024 ** 3)
            used_gb = tot_gb * (used_pct / 100)
            free_gb = tot_gb - used_gb
            free_str = fmt_gb(free_gb * (1024 ** 3))
            used_str = fmt_gb(used_gb * (1024 ** 3))
    elif vital_type == "CPU":
        if used_pct is not None:
            used_str = f"{used_pct:.2f}%"
            free_str = f"{100 - used_pct:.2f}%"

    return [
        ts.strftime("%d:%B:%Y %H:%M:%S"),
        year_val,
        month_val,
        day_val,
        weekday_val,
        alert,
        "CA",
        inst,
        r["job"],
        r["group"],
        sev,
        vital_type,
        cpu_str,
        mem_total_str if vital_type == "Memory" else "",
        disk_total_str if vital_type == "Disk" else "",
        volume if vital_type == "Disk" else "",
        free_str,
        used_str,
        summ,
        desc
    ]


def main():
    hw_specs = get_hardware_specs(PROM_URL)
    prom_assets, prom_meta = get_prometheus_assets_and_metadata(PROM_URL)
    inv_specs = get_inventory_specs(INVENTORY_FILE)

    output_csv = os.path.join(OUTPUT_DIR, OUTPUT_FILENAME)
    upload_csv = os.path.join(UPLOAD_DIR, OUTPUT_FILENAME)

    print(f"📖 Streaming log file: {LOG_FILE}")
    records = []
    with open(LOG_FILE, "r", encoding="utf-8", errors="ignore") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            try:
                a = json.loads(line)
            except Exception:
                continue

            ts_str = a.get("timestamp", "")
            if not ts_str:
                continue
            try:
                dt_obj = datetime.fromisoformat(ts_str.replace("Z", "+00:00"))
                ts = dt_obj if dt_obj.tzinfo else dt_obj.replace(tzinfo=timezone.utc)
            except Exception:
                continue

            if ts < START_DATE or ts > END_DATE:
                continue

            inst = a.get("instance", "").split(":")[0]
            desc = a.get("description", "")
            summ = a.get("summary", "")
            full_text = f"{summ}\n{desc}"

            asset = (
                extract_field(full_text, "asset")
                or a.get("asset")
                or a.get("Asset")
                or a.get("labels", {}).get("asset")
                or a.get("labels", {}).get("Asset")
                or prom_assets.get(inst, "")
            )

            if not asset or asset.strip().upper() != "CA":
                continue

            alert = a.get("alertname", "")
            sev = a.get("severity", "N/A")
            grp = extract_group(a, default_grp=prom_meta.get(inst, {}).get("group") or inv_specs.get(inst, {}).get("group") or "N/A")
            job = (
                extract_field(full_text, "job")
                or a.get("job")
                or a.get("labels", {}).get("job")
                or prom_meta.get(inst, {}).get("job")
                or inv_specs.get(inst, {}).get("job")
                or "alertmanager"
            )
            vital_extracted = extract_field(full_text, "company") or "SMC"

            records.append({
                "ts": ts,
                "alert": alert,
                "asset": "CA",
                "instance": inst,
                "job": job,
                "group": grp,
                "severity": sev,
                "vital_extracted": vital_extracted,
                "desc": desc,
                "summary": summ
            })

    records.sort(key=lambda x: x["ts"])
    print(f"✅ Matched {len(records)} records. Writing directly to {output_csv}...")

    with open(output_csv, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(HEADERS)
        for r in records:
            writer.writerow(format_row(r, hw_specs, inv_specs))

    print(f"🎉 CSV report created: {output_csv} ({os.path.getsize(output_csv) / (1024*1024):.2f} MB)")
    os.makedirs(UPLOAD_DIR, exist_ok=True)
    shutil.copyfile(output_csv, upload_csv)
    print(f"🔄 Synced to {upload_csv}")


if __name__ == "__main__":
    main()
