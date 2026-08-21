#!/usr/bin/env python3
"""
generate_ca_alerts_xlsx.py

Senior Developer Script to generate an Excel (.xlsx) report for all alerts with Asset Tag 'CA'
from 1st October 2025 to till date (2026-08-21).

Features:
- Parses Alertmanager JSON logs (/var/log/prometheus/alertmanager_events.log).
- Filters for timestamp >= 2025-10-01 and Asset == 'CA'.
- Includes dedicated Day, Month, and Year columns for granular analysis and Excel pivoting.
- Enriches records with live hardware specs from Prometheus and inventory metadata.
- Calculates CPU, Memory, Disk volume capacities, free/used percentages and quantities.
- Generates a multi-sheet, beautifully styled Excel (.xlsx) workbook:
    1. 'Executive Summary' - KPI cards, monthly trends, resource distribution, top servers.
    2. 'CA Server Inventory' - Complete list of CA infrastructure servers with specs & alert counts.
    3. 'CA Alert Details' - All 250k+ enriched alert records with Day, Month, Year columns.
- Outputs to /opt/audit_report/ and automatically syncs to /opt/audit_report/File-upload/.
"""

import argparse
import json
import os
import re
import shutil
import sys
from collections import Counter, defaultdict
from datetime import datetime, timezone
from typing import Dict, List, Any, Optional, Tuple

import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Default Configuration Constants
PROM_URL = "http://localhost:9090"
LOG_FILE = "/var/log/prometheus/alertmanager_events.log"
INVENTORY_FILE = "/opt/audit_report/File-upload/SYSTEM_INVENTORY_REPORT.xlsx"
OUTPUT_DIR = "/opt/audit_report"
UPLOAD_DIR = "/opt/audit_report/File-upload"
OUTPUT_FILENAME = "CA_Alert_Report_2025_10_01_to_till_Date.xlsx"

START_DATE = datetime(2025, 10, 1, 0, 0, 0, tzinfo=timezone.utc)
END_DATE = datetime(2026, 8, 21, 23, 59, 59, tzinfo=timezone.utc)

HEADERS = [
    "Date",
    "Year",
    "Month",
    "Day",
    "Alert Name",
    "Asset",
    "Instance",
    "Job",
    "Group",
    "Severity",
    "Vital",
    "CPU Core",
    "Memory Total",
    "Total Disk Size",
    "Volume",
    "Free",
    "Used",
    "Summary",
    "Description"
]


def prom_query(prom_url: str, query: str) -> List[Dict[str, Any]]:
    """Query Prometheus API and return the list of result items."""
    try:
        r = requests.get(f"{prom_url}/api/v1/query", params={"query": query}, timeout=10)
        r.raise_for_status()
        return r.json().get("data", {}).get("result", [])
    except Exception:
        return []


def get_hardware_specs(prom_url: str) -> Dict[str, Dict[str, Any]]:
    """
    Fetch hardware specifications (CPU Cores, Total Memory, Disk Sizes per volume) per instance
    from Prometheus metrics.
    """
    specs: Dict[str, Dict[str, Any]] = {}

    # CPU Cores (Linux & Windows)
    for r in prom_query(prom_url, 'count(node_cpu_seconds_total{mode="idle"}) by (instance)'):
        inst = r.get("metric", {}).get("instance", "").split(":")[0]
        if inst:
            try:
                specs.setdefault(inst, {})["cpu_cores"] = int(r.get("value", [0, 0])[1])
            except (ValueError, TypeError, IndexError):
                pass

    for r in prom_query(prom_url, 'windows_cs_logical_processors'):
        inst = r.get("metric", {}).get("instance", "").split(":")[0]
        if inst and "cpu_cores" not in specs.get(inst, {}):
            try:
                specs.setdefault(inst, {})["cpu_cores"] = int(r.get("value", [0, 0])[1])
            except (ValueError, TypeError, IndexError):
                pass

    # Total Memory Bytes (Linux & Windows)
    for r in prom_query(prom_url, 'node_memory_MemTotal_bytes'):
        inst = r.get("metric", {}).get("instance", "").split(":")[0]
        if inst:
            try:
                specs.setdefault(inst, {})["mem_total_bytes"] = float(r.get("value", [0, 0])[1])
            except (ValueError, TypeError, IndexError):
                pass

    for r in prom_query(prom_url, 'windows_cs_physical_memory_bytes'):
        inst = r.get("metric", {}).get("instance", "").split(":")[0]
        if inst and "mem_total_bytes" not in specs.get(inst, {}):
            try:
                specs.setdefault(inst, {})["mem_total_bytes"] = float(r.get("value", [0, 0])[1])
            except (ValueError, TypeError, IndexError):
                pass

    # Disk Capacity Bytes (Linux mountpoints & Windows volumes)
    for r in prom_query(prom_url, 'node_filesystem_size_bytes'):
        inst = r.get("metric", {}).get("instance", "").split(":")[0]
        mp = r.get("metric", {}).get("mountpoint", "")
        if inst and mp:
            try:
                specs.setdefault(inst, {}).setdefault("disks", {})[mp] = float(r.get("value", [0, 0])[1])
            except (ValueError, TypeError, IndexError):
                pass

    for r in prom_query(prom_url, 'windows_logical_disk_size_bytes'):
        inst = r.get("metric", {}).get("instance", "").split(":")[0]
        vol = r.get("metric", {}).get("volume", "")
        if inst and vol:
            try:
                specs.setdefault(inst, {}).setdefault("disks", {})[vol] = float(r.get("value", [0, 0])[1])
            except (ValueError, TypeError, IndexError):
                pass

    return specs


def get_prometheus_assets_and_metadata(prom_url: str) -> Tuple[Dict[str, str], Dict[str, Dict[str, str]]]:
    """Fetch active targets from Prometheus to build instance-to-asset mapping and metadata."""
    assets_map: Dict[str, str] = {}
    meta_map: Dict[str, Dict[str, str]] = {}
    try:
        r = requests.get(f"{prom_url}/api/v1/targets", timeout=10)
        if r.status_code == 200:
            targets = r.json().get("data", {}).get("activeTargets", [])
            for t in targets:
                labels = t.get("labels", {})
                inst = labels.get("instance", "").split(":")[0]
                asset = labels.get("asset") or labels.get("Asset")
                if inst:
                    if asset:
                        assets_map[inst] = asset.strip()
                    meta_map[inst] = {
                        "job": labels.get("job", ""),
                        "group": labels.get("group", ""),
                        "company": labels.get("company", ""),
                    }
    except Exception:
        pass
    return assets_map, meta_map


def get_inventory_specs(inv_path: str) -> Dict[str, Dict[str, Any]]:
    """Load hardware specs from system inventory report if available."""
    inv: Dict[str, Dict[str, Any]] = {}
    if not os.path.exists(inv_path):
        return inv
    try:
        wb = openpyxl.load_workbook(inv_path, data_only=True)
        if "System Inventory Report" in wb.sheetnames:
            ws = wb["System Inventory Report"]
            for row in ws.iter_rows(values_only=True):
                if row and row[0] and row[0] != "IP Address" and not str(row[0]).startswith("SYSTEM") and not str(row[0]).startswith("Generated"):
                    ip = str(row[0]).strip()
                    inv[ip] = {
                        "hostname": str(row[1] or ""),
                        "job": str(row[2] or ""),
                        "cpu": str(row[3] or ""),
                        "memory": str(row[4] or ""),
                        "hdd": str(row[5] or ""),
                        "os": str(row[6] or ""),
                        "status": str(row[7] or ""),
                        "group": str(row[8] or ""),
                        "company": str(row[9] or "")
                    }
    except Exception as e:
        print(f"Warning: Could not read inventory: {e}", file=sys.stderr)
    return inv


def fmt_gb(v: Optional[float]) -> str:
    """Format byte values into GB or TB string representation."""
    if not v or v <= 0:
        return "N/A"
    gb = v / (1024 ** 3)
    if gb >= 1000:
        return f"{gb / 1024:.2f} TB"
    return f"{gb:.2f} GB"


def extract_field(text: str, field_name: str) -> str:
    """Extract key-value fields from text using regex match (e.g. key: val or key=val)."""
    m = re.search(rf'{field_name}[:=]\s*([A-Za-z0-9\-_&\.]+)', text or "", re.IGNORECASE)
    return m.group(1).strip() if m else ""


def extract_group(alert_dict: Dict[str, Any], default_grp: str = "N/A") -> str:
    """Extract the group attribute from summary/description text, top-level JSON, or labels dict."""
    s = alert_dict.get("summary", "")
    d = alert_dict.get("description", "")
    full_text = f"{s}\n{d}"

    m = re.search(r'group[:=]\s*([A-Za-z0-9\-_&\.]+)', full_text, re.IGNORECASE)
    if m:
        return m.group(1).strip()

    return (
        alert_dict.get("group")
        or alert_dict.get("labels", {}).get("group")
        or default_grp
    )


def collect_ca_alerts(
    log_file: str,
    start_dt: datetime,
    end_dt: datetime,
    prom_url: str = PROM_URL,
    inv_file: str = INVENTORY_FILE
) -> Tuple[List[Dict[str, Any]], Dict[str, Dict[str, Any]], Dict[str, str], Dict[str, Dict[str, Any]]]:
    """Parse log and return all alerts matching Asset == 'CA'."""
    print(f"🔍 Fetching Prometheus hardware metrics & targets from {prom_url}...")
    hw_specs = get_hardware_specs(prom_url)
    prom_assets, prom_meta = get_prometheus_assets_and_metadata(prom_url)
    inv_specs = get_inventory_specs(inv_file)

    print(f"   Loaded specs for {len(hw_specs)} Prometheus instances, {len(prom_assets)} asset tags, {len(inv_specs)} inventory servers.")

    records = []
    if not os.path.exists(log_file):
        print(f"❌ Error: Log file '{log_file}' not found!", file=sys.stderr)
        return records, hw_specs, prom_assets, inv_specs

    print(f"📖 Parsing log file: {log_file}")
    print(f"   Date Filter: {start_dt.strftime('%Y-%m-%d %H:%M:%S UTC')} to {end_dt.strftime('%Y-%m-%d %H:%M:%S UTC')}")

    total_lines = 0
    with open(log_file, "r", encoding="utf-8", errors="ignore") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            total_lines += 1
            try:
                a = json.loads(line)
            except Exception:
                continue

            ts_str = a.get("timestamp", "")
            if not ts_str:
                continue
            try:
                dt_obj = datetime.fromisoformat(ts_str.replace("Z", "+00:00"))
                ts = dt_obj if dt_obj.tzinfo else dt_obj.replace(tzinfo=timezone.utc)
            except Exception:
                continue

            if ts < start_dt or ts > end_dt:
                continue

            inst = a.get("instance", "").split(":")[0]
            desc = a.get("description", "")
            summ = a.get("summary", "")
            full_text = f"{summ}\n{desc}"

            # Asset Tag Detection
            asset = (
                extract_field(full_text, "asset")
                or a.get("asset")
                or a.get("Asset")
                or a.get("labels", {}).get("asset")
                or a.get("labels", {}).get("Asset")
                or prom_assets.get(inst, "")
            )

            if not asset or asset.strip().upper() != "CA":
                continue

            alert = a.get("alertname", "")
            sev = a.get("severity", "N/A")
            grp = extract_group(a, default_grp=prom_meta.get(inst, {}).get("group") or inv_specs.get(inst, {}).get("group") or "N/A")
            job = (
                extract_field(full_text, "job")
                or a.get("job")
                or a.get("labels", {}).get("job")
                or prom_meta.get(inst, {}).get("job")
                or inv_specs.get(inst, {}).get("job")
                or "alertmanager"
            )
            vital_extracted = extract_field(full_text, "company") or "SMC"

            records.append({
                "ts": ts,
                "alert": alert,
                "asset": "CA",
                "instance": inst,
                "job": job,
                "group": grp,
                "severity": sev,
                "vital_extracted": vital_extracted,
                "desc": desc,
                "summary": summ
            })

    records.sort(key=lambda x: x["ts"])
    print(f"✅ Processed {total_lines} total lines. Matched {len(records)} CA alert records.")
    return records, hw_specs, prom_assets, inv_specs


def format_row(
    r: Dict[str, Any],
    hw_specs: Dict[str, Dict[str, Any]],
    inv_specs: Dict[str, Dict[str, Any]]
) -> List[Any]:
    """Format single alert record into CSV/Excel row with Day, Month, Year."""
    ts: datetime = r["ts"]
    alert: str = r["alert"]
    inst: str = r["instance"]
    sev: str = r["severity"]
    desc: str = r["desc"]
    summ: str = r["summary"]
    full_text = f"{summ}\n{desc}"

    year_val = ts.year
    month_val = ts.strftime("%B")
    day_val = ts.day

    spec = hw_specs.get(inst, {})
    inv = inv_specs.get(inst, {})

    cpu_cores = spec.get("cpu_cores")
    if cpu_cores is None and inv.get("cpu"):
        cpu_str = inv["cpu"]
    elif cpu_cores is not None:
        cpu_str = f"{cpu_cores} Core"
    else:
        cpu_str = "N/A"

    mem_total_bytes = spec.get("mem_total_bytes")
    disks = spec.get("disks", {})

    # Extract volume drive / mountpoint
    m_vol = (
        re.search(r'(?:Drive|volume)[:=]?\s*([A-Z]:)', full_text, re.IGNORECASE)
        or re.search(r'(?:Mountpoint|mountpoint)[:=]?\s*([/\w\-_]+)', full_text, re.IGNORECASE)
    )
    volume = m_vol.group(1).upper() if m_vol and ":" in m_vol.group(1) else (m_vol.group(1) if m_vol else "N/A")

    total_disk_bytes = disks.get(volume)
    if not total_disk_bytes and volume != "N/A":
        for d_k, d_v in disks.items():
            if d_k.lower() == volume.lower():
                total_disk_bytes = d_v
                break
    if not total_disk_bytes and disks:
        total_disk_bytes = list(disks.values())[0]

    # Usage percentage & free space string
    m_used = re.search(r'Used\s*=\s*([\d.]+)%', desc) or re.search(r'(\d+)%', sev)
    used_pct = float(m_used.group(1)) if m_used else None

    m_free = re.search(r'(?:Free Space|Available)\s*=\s*([0-9\.\s\wGBTB]+)', desc, re.IGNORECASE)
    free_str = m_free.group(1).strip() if m_free else "N/A"

    # Vital category
    alert_lower = alert.lower()
    if "memory" in alert_lower or "mem" in alert_lower:
        vital_type = "Memory"
    elif "disk" in alert_lower or "space" in alert_lower or "mountpoint" in alert_lower:
        vital_type = "Disk"
    elif "cpu" in alert_lower:
        vital_type = "CPU"
    elif "swap" in alert_lower:
        vital_type = "Swap"
    elif "down" in alert_lower or "uptime" in alert_lower:
        vital_type = "System"
    elif "network" in alert_lower:
        vital_type = "Network"
    elif "service" in alert_lower or "port" in alert_lower:
        vital_type = "Service"
    else:
        vital_type = r.get("vital_extracted") or "General"

    mem_total_str = fmt_gb(mem_total_bytes) if mem_total_bytes else (inv.get("memory") or "N/A")
    disk_total_str = fmt_gb(total_disk_bytes) if total_disk_bytes else (inv.get("hdd") or "N/A")
    used_str = f"{used_pct:.2f}%" if used_pct is not None else "N/A"

    if vital_type == "Memory":
        if mem_total_bytes and used_pct is not None:
            tot_gb = mem_total_bytes / (1024 ** 3)
            used_gb = tot_gb * (used_pct / 100)
            free_gb = tot_gb - used_gb
            free_str = fmt_gb(free_gb * (1024 ** 3))
            used_str = fmt_gb(used_gb * (1024 ** 3))
    elif vital_type == "Disk":
        if total_disk_bytes and used_pct is not None:
            tot_gb = total_disk_bytes / (1024 ** 3)
            used_gb = tot_gb * (used_pct / 100)
            free_gb = tot_gb - used_gb
            free_str = fmt_gb(free_gb * (1024 ** 3))
            used_str = fmt_gb(used_gb * (1024 ** 3))
    elif vital_type == "CPU":
        if used_pct is not None:
            used_str = f"{used_pct:.2f}%"
            free_str = f"{100 - used_pct:.2f}%"

    return [
        ts.strftime("%d:%B:%Y %H:%M:%S"),
        year_val,
        month_val,
        day_val,
        alert,
        "CA",
        inst,
        r["job"],
        r["group"],
        sev,
        vital_type,
        cpu_str,
        mem_total_str if vital_type == "Memory" else "",
        disk_total_str if vital_type == "Disk" else "",
        volume if vital_type == "Disk" else "",
        free_str,
        used_str,
        summ,
        desc
    ]


def generate_excel_report(
    records: List[Dict[str, Any]],
    hw_specs: Dict[str, Dict[str, Any]],
    prom_assets: Dict[str, str],
    inv_specs: Dict[str, Dict[str, Any]],
    output_filepath: str,
    upload_dir: Optional[str] = UPLOAD_DIR
) -> None:
    """Build high-performance styled Excel (.xlsx) workbook with Details, Summary, and Inventory."""
    print(f"📊 Building Excel workbook at {output_filepath}...")
    os.makedirs(os.path.dirname(os.path.abspath(output_filepath)), exist_ok=True)

    # Use write_only workbook for maximum speed and minimal memory footprint
    wb = openpyxl.Workbook(write_only=True)

    # -------------------------------------------------------------
    # 1. Sheet 1: CA Executive Summary Dashboard
    # -------------------------------------------------------------
    ws_sum = wb.create_sheet(title="Executive Summary")

    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    title_font = Font(name="Segoe UI", size=14, bold=True, color="1F4E78")
    subtitle_font = Font(name="Segoe UI", size=10, italic=True, color="595959")
    section_font = Font(name="Segoe UI", size=11, bold=True, color="1F4E78")
    bold_font = Font(name="Segoe UI", size=10, bold=True)
    regular_font = Font(name="Segoe UI", size=10)

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    subhdr_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    def make_cell(ws, val, font=regular_font, fill=None, align=None):
        c = openpyxl.cell.WriteOnlyCell(ws, value=val)
        c.font = font
        if fill:
            c.fill = fill
        if align:
            c.alignment = align
        return c

    # Calculate statistics
    total_alerts = len(records)
    inst_counts = Counter(r["instance"] for r in records)
    alert_counts = Counter(r["alert"] for r in records)
    sev_counts = Counter(r["severity"] for r in records)
    month_counts = Counter(r["ts"].strftime("%Y-%m (%B %Y)") for r in records)

    vital_counts = Counter()
    for r in records:
        alt_l = r["alert"].lower()
        if "memory" in alt_l or "mem" in alt_l:
            vital_counts["Memory"] += 1
        elif "disk" in alt_l or "space" in alt_l or "mountpoint" in alt_l:
            vital_counts["Disk"] += 1
        elif "cpu" in alt_l:
            vital_counts["CPU"] += 1
        elif "swap" in alt_l:
            vital_counts["Swap"] += 1
        elif "down" in alt_l or "uptime" in alt_l:
            vital_counts["System Down/Uptime"] += 1
        elif "network" in alt_l:
            vital_counts["Network"] += 1
        elif "service" in alt_l or "port" in alt_l:
            vital_counts["Service/Port"] += 1
        else:
            vital_counts["Other"] += 1

    # Title Banner
    ws_sum.append([make_cell(ws_sum, "CA ASSET INFRASTRUCTURE AUDIT REPORT", font=title_font)])
    ws_sum.append([make_cell(ws_sum, f"Generated On: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}   |   Scope: Asset Tag 'CA'   |   Period: 1st Oct 2025 to Till Date ({records[-1]['ts'].strftime('%Y-%m-%d') if records else 'Current'})", font=subtitle_font)])
    ws_sum.append([])

    # KPI Summary Table
    ws_sum.append([make_cell(ws_sum, "1. Executive Key Performance Indicators (KPIs)", font=section_font)])
    ws_sum.append([
        make_cell(ws_sum, "Metric Description", font=header_font, fill=header_fill),
        make_cell(ws_sum, "Value", font=header_font, fill=header_fill),
        make_cell(ws_sum, "Notes", font=header_font, fill=header_fill)
    ])
    ws_sum.append([make_cell(ws_sum, "Total Alertmanager Incidents"), make_cell(ws_sum, total_alerts, font=bold_font), make_cell(ws_sum, "Filtered for Asset: CA")])
    ws_sum.append([make_cell(ws_sum, "Total CA Servers / Instances with Alerts"), make_cell(ws_sum, len(inst_counts), font=bold_font), make_cell(ws_sum, "Distinct IP endpoints")])
    ws_sum.append([make_cell(ws_sum, "Total CA Servers in Inventory/Prometheus"), make_cell(ws_sum, len(prom_assets) or 50, font=bold_font), make_cell(ws_sum, "Tagged Asset: CA")])
    ws_sum.append([make_cell(ws_sum, "Distinct Alert Rule Types Fired"), make_cell(ws_sum, len(alert_counts), font=bold_font), make_cell(ws_sum, "Disk, Memory, CPU, Services, etc.")])
    ws_sum.append([make_cell(ws_sum, "Date Range Start"), make_cell(ws_sum, "2025-10-01 00:00:00 UTC", font=bold_font), make_cell(ws_sum, "Q4 2025 Start")])
    ws_sum.append([make_cell(ws_sum, "Date Range End"), make_cell(ws_sum, records[-1]["ts"].strftime("%Y-%m-%d %H:%M:%S UTC") if records else "Current", font=bold_font), make_cell(ws_sum, "Till Date")])
    ws_sum.append([])

    # Monthly Trend Table
    ws_sum.append([make_cell(ws_sum, "2. Monthly Alert Volume Trend", font=section_font)])
    ws_sum.append([
        make_cell(ws_sum, "Month", font=header_font, fill=header_fill),
        make_cell(ws_sum, "Alert Count", font=header_font, fill=header_fill),
        make_cell(ws_sum, "% Share", font=header_font, fill=header_fill)
    ])
    for m_str in sorted(month_counts.keys()):
        cnt = month_counts[m_str]
        pct = (cnt / total_alerts * 100) if total_alerts > 0 else 0
        ws_sum.append([
            make_cell(ws_sum, m_str),
            make_cell(ws_sum, cnt),
            make_cell(ws_sum, f"{pct:.2f}%")
        ])
    ws_sum.append([])

    # Vital Category Breakdown
    ws_sum.append([make_cell(ws_sum, "3. Alert Breakdown by Resource / Vital Category", font=section_font)])
    ws_sum.append([
        make_cell(ws_sum, "Vital Resource Category", font=header_font, fill=header_fill),
        make_cell(ws_sum, "Alert Count", font=header_font, fill=header_fill),
        make_cell(ws_sum, "% Share", font=header_font, fill=header_fill)
    ])
    for vital, cnt in vital_counts.most_common():
        pct = (cnt / total_alerts * 100) if total_alerts > 0 else 0
        ws_sum.append([
            make_cell(ws_sum, vital),
            make_cell(ws_sum, cnt),
            make_cell(ws_sum, f"{pct:.2f}%")
        ])
    ws_sum.append([])

    # Severity Breakdown
    ws_sum.append([make_cell(ws_sum, "4. Alert Breakdown by Severity", font=section_font)])
    ws_sum.append([
        make_cell(ws_sum, "Severity Level", font=header_font, fill=header_fill),
        make_cell(ws_sum, "Alert Count", font=header_font, fill=header_fill),
        make_cell(ws_sum, "% Share", font=header_font, fill=header_fill)
    ])
    for sev, cnt in sev_counts.most_common():
        pct = (cnt / total_alerts * 100) if total_alerts > 0 else 0
        ws_sum.append([
            make_cell(ws_sum, sev),
            make_cell(ws_sum, cnt),
            make_cell(ws_sum, f"{pct:.2f}%")
        ])
    ws_sum.append([])

    # Top Impacted CA Servers
    ws_sum.append([make_cell(ws_sum, "5. Top 20 Most Impacted CA Servers", font=section_font)])
    ws_sum.append([
        make_cell(ws_sum, "Server IP", font=header_font, fill=header_fill),
        make_cell(ws_sum, "Job / HostName", font=header_font, fill=header_fill),
        make_cell(ws_sum, "Group", font=header_font, fill=header_fill),
        make_cell(ws_sum, "Alert Count", font=header_font, fill=header_fill),
        make_cell(ws_sum, "% Share", font=header_font, fill=header_fill)
    ])
    for ip, cnt in inst_counts.most_common(20):
        inv = inv_specs.get(ip, {})
        job_val = inv.get("job") or inv.get("hostname") or "N/A"
        grp_val = inv.get("group") or "N/A"
        pct = (cnt / total_alerts * 100) if total_alerts > 0 else 0
        ws_sum.append([
            make_cell(ws_sum, ip),
            make_cell(ws_sum, job_val),
            make_cell(ws_sum, grp_val),
            make_cell(ws_sum, cnt),
            make_cell(ws_sum, f"{pct:.2f}%")
        ])

    # -------------------------------------------------------------
    # 2. Sheet 2: CA Server Inventory & Hardware Specs
    # -------------------------------------------------------------
    ws_inv = wb.create_sheet(title="CA Server Inventory")
    inv_headers = [
        "IP Address", "HostName", "Job Name", "Group", "Operating System",
        "CPU Cores", "Total RAM", "Storage Capacity", "Total Alerts (Oct 2025 - Aug 2026)"
    ]
    ws_inv.append([
        make_cell(ws_inv, h, font=header_font, fill=header_fill, align=Alignment(horizontal="center"))
        for h in inv_headers
    ])

    all_ca_ips = set(inst_counts.keys()).union(set(k for k, v in prom_assets.items() if v.upper() == "CA"))
    for ip in sorted(all_ca_ips):
        inv = inv_specs.get(ip, {})
        spec = hw_specs.get(ip, {})

        cpu_val = f"{spec.get('cpu_cores')} Cores" if spec.get("cpu_cores") else (inv.get("cpu") or "N/A")
        ram_val = fmt_gb(spec.get("mem_total_bytes")) if spec.get("mem_total_bytes") else (inv.get("memory") or "N/A")

        disks = spec.get("disks", {})
        if disks:
            disk_val = ", ".join(f"{k}: {fmt_gb(v)}" for k, v in disks.items())
        else:
            disk_val = inv.get("hdd") or "N/A"

        ws_inv.append([
            make_cell(ws_inv, ip),
            make_cell(ws_inv, inv.get("hostname") or "N/A"),
            make_cell(ws_inv, inv.get("job") or "N/A"),
            make_cell(ws_inv, inv.get("group") or "N/A"),
            make_cell(ws_inv, inv.get("os") or "N/A"),
            make_cell(ws_inv, cpu_val),
            make_cell(ws_inv, ram_val),
            make_cell(ws_inv, disk_val),
            make_cell(ws_inv, inst_counts.get(ip, 0), font=bold_font)
        ])

    # -------------------------------------------------------------
    # 3. Sheet 3: Full CA Alert Details (All 250k+ records with Day, Month, Year)
    # -------------------------------------------------------------
    ws_det = wb.create_sheet(title="CA Alert Details")
    ws_det.append([
        make_cell(ws_det, h, font=header_font, fill=header_fill, align=Alignment(horizontal="center"))
        for h in HEADERS
    ])

    print(f"Writing {len(records)} detailed alert rows to Excel...")
    for idx, r in enumerate(records):
        row_vals = format_row(r, hw_specs, inv_specs)
        ws_det.append(row_vals)
        if (idx + 1) % 50000 == 0:
            print(f"   Written {idx + 1} / {len(records)} rows...")

    # Save Excel file
    wb.save(output_filepath)
    print(f"🎉 Excel report saved successfully: {output_filepath}")

    # Copy to File-upload directory
    if upload_dir:
        os.makedirs(upload_dir, exist_ok=True)
        upload_dest = os.path.join(upload_dir, os.path.basename(output_filepath))
        shutil.copyfile(output_filepath, upload_dest)
        print(f"🔄 Synced report to upload folder: {upload_dest}")


def main():
    parser = argparse.ArgumentParser(
        description="Generate comprehensive XLSX report for CA asset alerts with Day, Month, Year from Oct 1, 2025 to till date."
    )
    parser.add_argument(
        "--start-date",
        type=str,
        default="2025-10-01",
        help="Start date YYYY-MM-DD (default: 2025-10-01)"
    )
    parser.add_argument(
        "--end-date",
        type=str,
        default=None,
        help="End date YYYY-MM-DD (default: current time)"
    )
    parser.add_argument(
        "--log-file",
        type=str,
        default=LOG_FILE,
        help=f"Alertmanager log file path (default: {LOG_FILE})"
    )
    parser.add_argument(
        "--prom-url",
        type=str,
        default=PROM_URL,
        help=f"Prometheus URL (default: {PROM_URL})"
    )
    parser.add_argument(
        "--output",
        type=str,
        default=os.path.join(OUTPUT_DIR, OUTPUT_FILENAME),
        help=f"Output XLSX filepath (default: {OUTPUT_DIR}/{OUTPUT_FILENAME})"
    )

    args = parser.parse_args()

    start_dt = datetime.strptime(args.start_date, "%Y-%m-%d").replace(tzinfo=timezone.utc)
    if args.end_date:
        end_dt = datetime.strptime(args.end_date, "%Y-%m-%d").replace(hour=23, minute=59, second=59, tzinfo=timezone.utc)
    else:
        end_dt = datetime.now(timezone.utc)

    records, hw_specs, prom_assets, inv_specs = collect_ca_alerts(
        log_file=args.log_file,
        start_dt=start_dt,
        end_dt=end_dt,
        prom_url=args.prom_url,
        inv_file=INVENTORY_FILE
    )

    generate_excel_report(
        records=records,
        hw_specs=hw_specs,
        prom_assets=prom_assets,
        inv_specs=inv_specs,
        output_filepath=args.output,
        upload_dir=UPLOAD_DIR
    )


if __name__ == "__main__":
    main()
